from flask import Blueprint,Response, render_template, session, redirect, url_for, request, jsonify, current_app,send_from_directory

from config import Config

from functools import wraps
from routes.utils import abort
from models.proyecto import Proyecto
from models.cloud_ejecucion import CloudEjecucion
from cryptography.fernet import Fernet
import mysql.connector
import boto3
import csv
import io
import re
from botocore.exceptions import ClientError, EndpointConnectionError
import importlib
from db import get_db_connection  # asegurate de tenerlo arriba
from datetime import datetime
from redis import Redis
from rq import Queue
from tasks.cloud.aws import discovery_roles_job

import re
from datetime import datetime
from io import BytesIO

from flask import Response, session
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font

from openpyxl import Workbook
from openpyxl.styles import Alignment
from io import BytesIO

proyecto_bp = Blueprint('proyecto', __name__)


# ------------------------------------------------------------------
# NO CACHE
# ------------------------------------------------------------------
@proyecto_bp.after_request
def no_cache_dashboard(response):
    response.headers['Cache-Control'] = 'no-store'
    return response


# ------------------------------------------------------------------
# LOGIN REQUIRED
# ------------------------------------------------------------------
def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return wrapper


# ------------------------------------------------------------------
# LISTADO PROYECTOS
# ------------------------------------------------------------------
@proyecto_bp.route('/proyectos')
@login_required
def index():
    sector_id = session.get('sector_id')
    proyectos = Proyecto.get_proyectos(sector_id)
    tipos_servicio = Proyecto.get_tipos_servicio()

    return render_template(
        'inicio/proyectos.html',
        proyectos=proyectos,
        tipos_servicio=tipos_servicio
    )


# ------------------------------------------------------------------
# CREAR PROYECTO
# ------------------------------------------------------------------
@proyecto_bp.route('/proyecto/crear', methods=['POST'])
@login_required
def crear_proyecto():

    sector_id = session.get('sector_id')
    usuario_creador_id = session.get('user_id')
    estado_id = 1

    titulo = request.form.get('titulo')
    cliente = request.form.get('cliente')
    tipo_proyecto = request.form.get('tipo_proyecto')
    tipo_servicio = request.form.get('tipo_servicio')
    autenticado = request.form.get('autenticado')

    if not titulo or not cliente:
        return jsonify({"success": False, "message": "Campos obligatorios"}), 400

    try:
        Proyecto.insert_proyecto(
            titulo=titulo,
            cliente=cliente,
            sector_id=sector_id,
            usuario_creador_id=usuario_creador_id,
            tipo_proyecto=tipo_proyecto,
            tipo_servicio=tipo_servicio,
            autenticado=autenticado,
            estado_id=estado_id
        )

        return jsonify({"success": True, "message": "Proyecto creado correctamente"})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


# ------------------------------------------------------------------
# DETALLE PROYECTO
# ------------------------------------------------------------------
@proyecto_bp.route('/proyecto/<int:proyecto_id>')
@login_required
def proyecto_detalle(proyecto_id):

    sector_id = session.get('sector_id')
    proyecto = Proyecto.get_by_id(proyecto_id, sector_id)

    if not proyecto:
        abort(404)

    return render_template(
        'proyecto/detalle.html',
        proyecto=proyecto
    )


# ------------------------------------------------------------------
# GUARDAR CONFIG CLOUD
# ------------------------------------------------------------------
@proyecto_bp.route('/proyecto/<int:proyecto_id>/cloud-config', methods=['POST'])
@login_required
def guardar_cloud_config(proyecto_id):

    sector_id = session.get('sector_id')
    proyecto = Proyecto.get_by_id(proyecto_id, sector_id)

    if not proyecto:
        return jsonify({
            "success": False,
            "message": "Proyecto no encontrado"
        }), 404

    if proyecto['tipo_proyecto'] != 'CLOUD':
        return jsonify({
            "success": False,
            "message": "No es proyecto Cloud"
        }), 400

    # 🔹 Campos del form
    auth_method = request.form.get('auth_method')
    access_key = request.form.get('access_key')
    secret_key = request.form.get('secret_key')
    arn_role = request.form.get('arn_role')
    external_id = request.form.get('external_id')
    region = request.form.get('region')

    if not auth_method or not region:
        return jsonify({
            "success": False,
            "message": "Método de autenticación y región son obligatorios"
        }), 400

    region = region.strip()
    aws_account_id = None
    secret_key_encrypted = None

    try:

        # ==========================================================
        # 🔐 MODO ROLE
        # ==========================================================
        if auth_method == "role":

            if not arn_role:
                return jsonify({
                    "success": False,
                    "message": "Role ARN es obligatorio"
                }), 400

            sts = boto3.client('sts')

            assume_params = {
                "RoleArn": arn_role.strip(),
                "RoleSessionName": "RedScopeValidationSession"
            }

            if external_id:
                assume_params["ExternalId"] = external_id.strip()

            assumed = sts.assume_role(**assume_params)
            creds = assumed["Credentials"]

            sts_temp = boto3.client(
                "sts",
                aws_access_key_id=creds["AccessKeyId"],
                aws_secret_access_key=creds["SecretAccessKey"],
                aws_session_token=creds["SessionToken"],
                region_name=region
            )

            identity = sts_temp.get_caller_identity()
            aws_account_id = identity.get("Account")

            # Limpiar claves si estaban enviadas
            access_key = None
            secret_key_encrypted = None

        # ==========================================================
        # 🔑 MODO ACCESS KEYS
        # ==========================================================
        elif auth_method == "keys":

            if not access_key or not secret_key:
                return jsonify({
                    "success": False,
                    "message": "Access Key y Secret Key son obligatorios"
                }), 400

            access_key = access_key.strip()

            sts = boto3.client(
                'sts',
                aws_access_key_id=access_key,
                aws_secret_access_key=secret_key,
                region_name=region
            )

            identity = sts.get_caller_identity()
            aws_account_id = identity.get("Account")

            # Cifrar secret_key
            fernet = Fernet(current_app.config['FERNET_KEY'])
            secret_key_encrypted = fernet.encrypt(secret_key.encode()).decode()

            arn_role = None
            external_id = None

        else:
            return jsonify({
                "success": False,
                "message": "Método de autenticación inválido"
            }), 400

    except EndpointConnectionError:
        return jsonify({
            "success": False,
            "message": "Región AWS inválida"
        }), 400

    except ClientError as e:
        return jsonify({
            "success": False,
            "message": f"Error AWS: {str(e)}"
        }), 400

    except Exception as e:
        print("AWS ERROR:", str(e))
        return jsonify({
            "success": False,
            "message": "Error validando credenciales AWS"
        }), 500

    # ==========================================================
    # 💾 Guardar en DB
    # ==========================================================
    try:

        Proyecto.guardar_cloud_config(
            proyecto_id=proyecto_id,
            auth_method=auth_method,
            access_key=access_key,
            secret_key=secret_key_encrypted,
            role_arn=arn_role,
            external_id=external_id,
            region=region,
            aws_account_id=aws_account_id
        )

        return jsonify({
            "success": True,
            "message": f"Configuración guardada correctamente (Account ID: {aws_account_id})"
        })

    except mysql.connector.Error as e:
        print("MYSQL ERROR:", e)
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

    except Exception as e:
        print("ERROR FINAL:", str(e))
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@proyecto_bp.route('/proyecto/<int:proyecto_id>/cloud')
@login_required
def proyecto_cloud_workspace(proyecto_id):

    sector_id = session.get('sector_id')
    proyecto = Proyecto.get_by_id(proyecto_id, sector_id)

    if not proyecto:
        abort(404)

    if proyecto['tipo_proyecto'] != 'CLOUD':
        abort(403)

    servicios_aws = Proyecto.get_servicios_aws_by_id(
        proyecto['tipo_servicio_id']
    )

    return render_template(
        'proyecto/proyectos-cloud/index.html',
        proyecto=proyecto,
        servicios_aws=servicios_aws
    )


@proyecto_bp.route('/cloud/acciones/<int:servicio_id>')
@login_required
def obtener_acciones_cloud(servicio_id):

    acciones = Proyecto.get_servicios_aws_acciones(servicio_id)

    return jsonify({
        "success": True,
        "acciones": acciones
    })


@proyecto_bp.route('/cloud/run-roles', methods=['POST'])
@login_required
def run_roles():
    data = request.get_json(silent=True) or {}

    proyecto_id = data.get('proyecto_id')
    accion_id = data.get('accion_id')
    usuario_id = session.get('user_id')

    if not proyecto_id or not accion_id:
        return jsonify({
            "success": False,
            "message": "proyecto_id y accion_id son obligatorios"
        }), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO cloud_ejecuciones
        (proyecto_id, accion_id, usuario_id, estado, fecha_creacion, estado_id)
        VALUES (%s,%s,%s,'QUEUED', NOW(), 1)
        ON DUPLICATE KEY UPDATE
            usuario_id = VALUES(usuario_id),
            estado = 'QUEUED',
            nivel_resultado = NULL,
            codigo_resultado = NULL,
            resultado = NULL,
            error = NULL,
            fecha_creacion = NOW(),
            fecha_fin = NULL,
            estado_id = 1
    """, (proyecto_id, accion_id, usuario_id))

    # Siempre obtenemos el ID correcto
    cursor.execute("""
        SELECT id FROM cloud_ejecuciones
        WHERE proyecto_id=%s AND accion_id=%s
    """, (proyecto_id, accion_id))

    ejecucion_id = cursor.fetchone()[0]

    conn.commit()
    cursor.close()
    conn.close()

    q = Queue(connection=Config.redis_conn)

    accion = Proyecto.get_accion_by_id(accion_id)
    if not accion:
        return jsonify({
            "success": False,
            "message": "Acción inválida"
        }), 400

    handler_path = accion['handler']
    module_path, function_name = handler_path.rsplit('.', 1)

    try:
        module = importlib.import_module(f"tasks.{module_path}")
        func = getattr(module, function_name)
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Error cargando handler: {str(e)}"
        }), 500

    q.enqueue(func, ejecucion_id, proyecto_id)

    return jsonify({
        "success": True,
        "ejecucion_id": ejecucion_id
    })


@proyecto_bp.route('/cloud/resultados/<int:proyecto_id>')
@login_required
def obtener_resultados_cloud(proyecto_id):

    sector_id = session.get('sector_id')
    proyecto = Proyecto.get_by_id(proyecto_id, sector_id)

    if not proyecto:
        return jsonify({"success": False}), 404

    data = Proyecto.get_data_ejecucion_cloud(proyecto_id)

    return jsonify({
        "success": True,
        "data": data
    })


@proyecto_bp.route('/proyecto/<int:proyecto_id>/cloud/ejecucion/<int:ejecucion_id>', methods=['GET'])
@login_required
def obtener_detalle_ejecucion(proyecto_id, ejecucion_id):

    sector_id = session.get('sector_id')
    proyecto = Proyecto.get_by_id(proyecto_id, sector_id)

    if not proyecto:
        abort(404)

    data = Proyecto.get_data_ejecuciones_para_analisis(
        proyecto_id,
        ejecucion_id
    )

    if not data:
        return jsonify({"error": "No encontrado"}), 404

    # TRAIGO LOS INFDINGS
    findings = CloudEjecucion.extract_interesting(data["resultado"])
    data["interesting"] = findings

    return jsonify(data)


@proyecto_bp.route('/proyecto/<int:proyecto_id>/cloud/ejecucion/<int:ejecucion_id>/hallazgos')
@login_required
def gestionar_hallazgos(proyecto_id, ejecucion_id):
    sector_id = session.get('sector_id')
    proyecto = Proyecto.get_by_id(proyecto_id, sector_id)
    if not proyecto:
        abort(404)
    data = Proyecto.get_data_ejecuciones_para_analisis(
        proyecto_id,
        ejecucion_id
    )
    if not data:
        abort(404)
    findings = CloudEjecucion.extract_interesting(data["resultado"])
    return render_template(
        'proyecto/proyectos-cloud/GestionHallazgos.html',
        proyecto=proyecto,
        ejecucion=data,
        findings=findings
    )


@proyecto_bp.route('/proyecto/security-rule/<check_id>')
@login_required
def gestionar_findings(check_id):

    data = Proyecto.get_security_rules(check_id)
    severidades = Proyecto.get_severidades()
    combo_findings = Proyecto.get_combo_estados_findings()  # devuelve lista de estados

    return jsonify({
        'success': True,
        'rule_exists': True if data else False,
        'data': data,
        'severidades': severidades,
        'combo_findings': combo_findings       # 🔹 agregamos los estados aquí
    })
    
@proyecto_bp.route('/proyecto/security-rule', methods=['POST'])
@login_required
def insert_security_rule():

    data = request.get_json()

    rule_id = Proyecto.insert_security_rule(data)

    return jsonify({
        "success": True,
        "rule_id": rule_id
    })
    

@proyecto_bp.route('/proyecto/finding', methods=['POST'])
@login_required
def insert_finding():
    data = request.get_json()
    usuario_id = session.get("user_id")
    finding_id = Proyecto.insert_finding(data, usuario_id)

    # nuevas evidencias
    if data.get("evidencias"):
        Proyecto.insert_evidences(finding_id, data["evidencias"])

    # eliminadas
    if data.get("evidencias_eliminadas"):
        Proyecto.delete_evidences(data["evidencias_eliminadas"])

    return jsonify({
        "success": True
    })
    

@proyecto_bp.route("/proyecto/finding/<int:finding_id>")
@login_required
def api_get_finding(finding_id):
    hallazgo = Proyecto.get_finding(finding_id)
    evidencias = Proyecto.get_finding_evidencias(finding_id)
    combo_findings = Proyecto.get_combo_estados_findings()

    return jsonify({
        "success": True,
        "hallazgo": hallazgo,
        "evidencias": [e["file_path"] for e in evidencias],
        "combo_findings": combo_findings
    })
    
@proyecto_bp.route('/proyecto/finding/<int:proyecto_id>/<string:check_id>', methods=['GET'])
def get_finding_by_check(proyecto_id, check_id):
    # Obtenemos el finding principal
    finding = Proyecto.get_finding(check_id=check_id, proyecto_id=proyecto_id)
    
    if not finding:
        return jsonify({
            "success": False,
            "message": "No se encontró el finding"
        }), 404

    # Obtenemos el finding_id desde el finding devuelto
    finding_id = finding['id']  # asumimos que get_finding devuelve un dict con 'id'

    # Traemos las evidencias asociadas
    evidencias_img = Proyecto.get_finding_evidencias_img(finding_id)

    return jsonify({
        "success": True,
        "data": {
            "finding": finding,
            "evidencias_img": evidencias_img
        }
    })
    
@proyecto_bp.route("/uploads/findings/<path:filename>")
@login_required
def serve_evidencias(filename):
    import os
    from flask import send_from_directory
    path = os.path.join(Config.BASE_DIR, "uploads", "findings")
    return send_from_directory(path, filename)


# *****************************  Reportes  ***************************
#------------- CSV ------------
@proyecto_bp.route('/proyecto/<int:proyecto_id>/export/xlsx')
@login_required
def exportar_xlsx(proyecto_id):
    sector_id = session.get('sector_id')
    proyecto = Proyecto.get_by_id(proyecto_id, sector_id)

    if not proyecto:
        abort(404)

    data = Proyecto.get_data_reporte_csv(proyecto_id)

    # -----------------------------
    # 🧠 Nombre dinámico PRO
    # -----------------------------
    def limpiar(texto):
        if not texto:
            return "sin_valor"
        return re.sub(r'[^a-zA-Z0-9_-]', '_', str(texto))

    if data:
        titulo = limpiar(data[0].get('proyecto_titulo'))
        proveedores = {row.get('proveedor', 'sin_proveedor') for row in data}
        proveedor = limpiar("_".join(proveedores))
    else:
        titulo = f"proyecto_{proyecto_id}"
        proveedor = "sin_proveedor"

    filename = f"{titulo}_{proveedor}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # -----------------------------
    # 📊 Excel
    # -----------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Findings"

    headers = [
        'proveedor','servicio','check_id','titulo','descripcion',
        'riesgo','condicion logica','remediacion','referencia',
        'resource_id','estado'
    ]

    ws.append(headers)

    # -----------------------------
    # 🎨 Estilos
    # -----------------------------
    wrap = Alignment(wrap_text=True, vertical="top")

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    severity_colors = {
        "INFORMATIVO": "808080",      # gray
        "BAJO": "00B050",       # green
        "MEDIO": "FFA500",    # orange
        "ALTO": "FF0000",      # red
        "CRITICO": "800080"   # purple
    }

    # Header style
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap

    # -----------------------------
    # 📥 Data
    # -----------------------------
    for row in data:
        fila = [
            row.get('proveedor', ''),
            row.get('servicio', ''),
            row.get('check_id', ''),
            row.get('titulo', ''),
            row.get('descripcion', ''),
            row.get('severidad', ''),
            row.get('condicion_logica', ''),
            row.get('remediacion', ''),
            row.get('referencia', ''),
            row.get('resource_id', ''),
            row.get('estado', '')
        ]

        ws.append(fila)
        current_row = ws.max_row

        for col_idx, value in enumerate(fila, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.alignment = wrap

            # 🎨 Color por severidad
            if col_idx == 6:  # columna severity
                sev = str(value).upper()
                if sev in severity_colors:
                    cell.fill = PatternFill(
                        start_color=severity_colors[sev],
                        end_color=severity_colors[sev],
                        fill_type="solid"
                    )

    # -----------------------------
    # 🔥 AutoFilter (filtros en columnas)
    # -----------------------------
    ws.auto_filter.ref = ws.dimensions

    # -----------------------------
    # 🔒 Freeze header
    # -----------------------------
    ws.freeze_panes = "A2"

    # -----------------------------
    # 📏 Auto width columnas
    # -----------------------------
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # -----------------------------
    # 💾 Guardar
    # -----------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),  # 🔥 bytes reales
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Cache-Control": "no-cache"
        }
    )